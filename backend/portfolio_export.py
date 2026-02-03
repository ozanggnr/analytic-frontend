# Add this new endpoint after the existing export endpoint

@app.get("/api/export/portfolio")
def export_portfolio(symbols: str, period: str = "daily"):
    """
    Export only specific portfolio symbols
    symbols: comma-separated list of symbols
    period: daily, weekly, monthly
    """
    if period not in ["daily", "weekly", "monthly"]:
        raise HTTPException(status_code=400, detail="Invalid period")
    
    symbol_list = symbols.split(',')
    
    # Analyze each symbol
    results = []
    for sym in symbol_list:
        sym = sym.strip()
        if not sym:
            continue
            
        analysis = analyze_stock(sym, period)
        if analysis:
            results.append(analysis)
    
    if not results:
        raise HTTPException(status_code=404, detail="No data for portfolio stocks")
    
    # Convert to Excel (same format as market export)
    import pandas as pd
    import io
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    
    df = pd.DataFrame(results)
    
    # Select columns
    cols = ['symbol', 'name', 'price', 'change_pct', 'ma_20', 'rsi', 
            'volatility', 'prediction', 'is_buyable']
    df = df[cols]
    
    # Rename for Excel
    df.columns = ['Symbol', 'Name', 'Price (₺)', 'Change %', 'MA(20)', 
                  'RSI', 'Volatility %', 'Analysis', 'Buy Signal']
    
    # Create Excel with styling
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Portfolio', index=False)
        
        ws = writer.sheets['Portfolio']
        
        # Header styling
        header_fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=portfolio_{period}.xlsx"}
    )
